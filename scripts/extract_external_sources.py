from __future__ import annotations

import argparse
import html
import json
import re
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def extract_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        rows: list[list[str]] = []
        non_empty = 0
        for row in ws.iter_rows(values_only=True):
            cleaned = [clean_text(cell) for cell in row]
            if any(cleaned):
                non_empty += 1
                rows.append(cleaned)

        max_cols = max((len(r) for r in rows), default=0)
        normalized = [r + [""] * (max_cols - len(r)) for r in rows]

        sheets.append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "non_empty_rows": non_empty,
                "rows": normalized,
                "preview_rows": normalized[:20],
            }
        )

    return {"path": str(path), "sheets": sheets}


def parse_mxgraph_html(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    match = re.search(r'data-mxgraph="([^"]+)"', text)
    if not match:
        raise RuntimeError("data-mxgraph attribute not found")

    raw_attr = html.unescape(match.group(1))
    config = json.loads(raw_attr)
    xml_text = html.unescape(config["xml"])
    root = ET.fromstring(xml_text)

    labels: list[dict[str, Any]] = []
    for cell in root.iter():
        if not cell.tag.endswith("mxCell"):
            continue
        value = clean_text(html.unescape(cell.attrib.get("value", "")))
        if not value:
            continue
        geometry = None
        for child in cell:
            if child.tag.endswith("mxGeometry"):
                geometry = {
                    "x": child.attrib.get("x"),
                    "y": child.attrib.get("y"),
                    "width": child.attrib.get("width"),
                    "height": child.attrib.get("height"),
                }
                break
        labels.append(
            {
                "id": cell.attrib.get("id"),
                "parent": cell.attrib.get("parent"),
                "edge": cell.attrib.get("edge"),
                "vertex": cell.attrib.get("vertex"),
                "value": value,
                "geometry": geometry,
            }
        )

    return {
        "path": str(path),
        "label_count": len(labels),
        "labels": labels,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--html", required=True)
    parser.add_argument("--outdir", required=True)
    args = parser.parse_args()

    outdir = Path(args.outdir)
    workbook_data = extract_workbook(Path(args.xlsx))
    html_data = parse_mxgraph_html(Path(args.html))

    write_json(outdir / "workbook_extract.json", workbook_data)
    write_json(outdir / "html_extract.json", html_data)

    print(f"workbook_sheets={len(workbook_data['sheets'])}")
    print(f"html_labels={html_data['label_count']}")
    print(f"outdir={outdir}")


if __name__ == "__main__":
    main()
