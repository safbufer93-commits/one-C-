from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def scalar(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def compact(value: Any, limit: int = 240) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r", "\\r").replace("\n", "\\n")
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def cell_fill(cell) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return None
    color = fill.fgColor
    if color is None:
        return fill.fill_type
    rgb = color.rgb or color.indexed or color.theme
    return f"{fill.fill_type}:{rgb}"


def sheet_stats(ws_formula, ws_values) -> dict[str, Any]:
    non_empty = []
    formulas = []
    constants = []
    fills = Counter()
    for row in ws_formula.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            coord = cell.coordinate
            cached = ws_values[coord].value
            fill = cell_fill(cell)
            if fill:
                fills[fill] += 1
            item = {
                "cell": coord,
                "row": cell.row,
                "col": cell.column,
                "value": scalar(value),
                "cached_value": scalar(cached),
                "number_format": cell.number_format,
                "fill": fill,
            }
            non_empty.append(item)
            if isinstance(value, str) and value.startswith("="):
                formulas.append(item)
            else:
                constants.append(item)

    preview_cells = sorted(
        non_empty,
        key=lambda x: (x["row"], x["col"]),
    )[:300]
    bottom_preview = sorted(
        non_empty,
        key=lambda x: (x["row"], x["col"]),
    )[-80:]

    formula_prefixes = Counter()
    for item in formulas:
        formula = str(item["value"])
        prefix = formula.split("(", 1)[0][:80]
        formula_prefixes[prefix] += 1

    return {
        "title": ws_formula.title,
        "state": ws_formula.sheet_state,
        "max_row": ws_formula.max_row,
        "max_column": ws_formula.max_column,
        "dimension": ws_formula.calculate_dimension(),
        "non_empty_count": len(non_empty),
        "formula_count": len(formulas),
        "constant_count": len(constants),
        "merged_ranges": [str(r) for r in ws_formula.merged_cells.ranges],
        "tables": list(ws_formula.tables.keys()),
        "auto_filter": str(ws_formula.auto_filter.ref) if ws_formula.auto_filter and ws_formula.auto_filter.ref else None,
        "freeze_panes": str(ws_formula.freeze_panes) if ws_formula.freeze_panes else None,
        "fills": fills.most_common(20),
        "formula_prefixes": formula_prefixes.most_common(30),
        "preview_cells": preview_cells,
        "bottom_preview_cells": bottom_preview,
        "formula_cells": sorted(formulas, key=lambda x: (x["row"], x["col"]))[:1200],
        "constant_cells": sorted(constants, key=lambda x: (x["row"], x["col"]))[:1200],
    }


def write_markdown(summary: dict[str, Any], path: Path) -> None:
    lines: list[str] = []
    lines.append("# Анализ Excel-калькулятора Volvo")
    lines.append("")
    lines.append(f"Файл: `{summary['path']}`")
    lines.append(f"Листов: {len(summary['sheets'])}")
    lines.append("")
    lines.append("## Листы")
    lines.append("")
    lines.append("| Лист | Состояние | Размер | Непустых | Формул | Констант |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    for sheet in summary["sheets"]:
        lines.append(
            f"| {sheet['title']} | {sheet['state']} | {sheet['dimension']} | "
            f"{sheet['non_empty_count']} | {sheet['formula_count']} | {sheet['constant_count']} |"
        )

    if summary["defined_names"]:
        lines.append("")
        lines.append("## Именованные диапазоны")
        lines.append("")
        for name in summary["defined_names"][:200]:
            lines.append(f"- `{name}`")

    for sheet in summary["sheets"]:
        lines.append("")
        lines.append(f"## Лист: {sheet['title']}")
        lines.append("")
        lines.append(
            f"Размер: `{sheet['dimension']}`, непустых: {sheet['non_empty_count']}, "
            f"формул: {sheet['formula_count']}, констант: {sheet['constant_count']}."
        )
        if sheet["merged_ranges"]:
            lines.append(f"Объединения: {', '.join(sheet['merged_ranges'][:30])}")
        if sheet["formula_prefixes"]:
            lines.append("")
            lines.append("Частые типы формул:")
            for prefix, count in sheet["formula_prefixes"][:15]:
                lines.append(f"- `{prefix}`: {count}")
        lines.append("")
        lines.append("Первые непустые ячейки:")
        lines.append("")
        lines.append("| Ячейка | Значение / формула | Кэш | Формат | Заливка |")
        lines.append("|---|---|---|---|---|")
        for item in sheet["preview_cells"][:120]:
            lines.append(
                f"| {item['cell']} | `{compact(item['value'])}` | `{compact(item['cached_value'])}` | "
                f"`{compact(item['number_format'], 80)}` | `{compact(item['fill'], 80)}` |"
            )

        if sheet["formula_cells"]:
            lines.append("")
            lines.append("Первые формулы:")
            lines.append("")
            lines.append("| Ячейка | Формула | Кэш |")
            lines.append("|---|---|---|")
            for item in sheet["formula_cells"][:120]:
                lines.append(
                    f"| {item['cell']} | `{compact(item['value'], 500)}` | `{compact(item['cached_value'])}` |"
                )

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    if len(sys.argv) < 3:
        print("usage: analyze_excel_workbook.py <xlsx> <out_dir>", file=sys.stderr)
        return 2

    workbook_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    wb_formula = load_workbook(workbook_path, data_only=False)
    wb_values = load_workbook(workbook_path, data_only=True)

    summary = {
        "path": str(workbook_path),
        "sheetnames": wb_formula.sheetnames,
        "active_sheet": wb_formula.active.title if wb_formula.active else None,
        "defined_names": sorted(str(name) for name in wb_formula.defined_names),
        "sheets": [],
    }

    for ws_formula in wb_formula.worksheets:
        ws_values = wb_values[ws_formula.title]
        summary["sheets"].append(sheet_stats(ws_formula, ws_values))

    json_path = out_dir / "volvo_calculator_analysis.json"
    md_path = out_dir / "volvo_calculator_analysis.md"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(summary, md_path)

    print(json.dumps({
        "json": str(json_path),
        "markdown": str(md_path),
        "sheets": summary["sheetnames"],
        "formula_total": sum(s["formula_count"] for s in summary["sheets"]),
        "non_empty_total": sum(s["non_empty_count"] for s in summary["sheets"]),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
