from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def short(value: Any, limit: int = 700) -> str:
    text = clean(value)
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def normalize_formula(formula: str) -> str:
    # Keep absolute setting rows ($BE$6) intact enough to understand, but collapse
    # row-specific copies for repeated line formulas.
    return re.sub(r"(?<!\$)([A-Z]{1,3})(\d+)", r"\1#", formula)


def cell_kind(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    return "input"


def unique_nonempty(values: list[Any], limit: int = 80) -> list[str]:
    result: list[str] = []
    seen = set()
    for value in values:
        text = clean(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
        if len(result) >= limit:
            break
    return result


def extract_strings_from_formula(formula: str) -> list[str]:
    return [x for x in re.findall(r'"([^"]+)"', formula or "") if x]


def main() -> int:
    if len(sys.argv) < 3:
        print("usage: summarize_volvo_calculator.py <xlsx> <out_dir>", file=sys.stderr)
        return 2

    workbook_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    wb_formula = load_workbook(workbook_path, data_only=False)
    wb_values = load_workbook(workbook_path, data_only=True)
    ws = wb_formula["Калькулятор"]
    ws_values = wb_values["Калькулятор"]
    aux = wb_formula["Сопутсвующие расходы"]
    aux_values = wb_values["Сопутсвующие расходы"]

    header_row = 5
    first_data_row = 6
    last_data_row = 1500
    max_col = ws.max_column

    columns: list[dict[str, Any]] = []
    all_delivery_strings = set()
    all_ip_strings = set()
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        header = clean(ws.cell(header_row, col).value)
        sample_formula_value = ws.cell(first_data_row, col).value
        sample_cached_value = ws_values.cell(first_data_row, col).value

        kinds = Counter()
        unique_inputs: list[Any] = []
        formula_templates = Counter()
        for row in range(first_data_row, last_data_row + 1):
            value = ws.cell(row, col).value
            kind = cell_kind(value)
            kinds[kind] += 1
            if kind == "input" and value is not None:
                unique_inputs.append(value)
            if kind == "formula":
                formula_templates[normalize_formula(str(value))] += 1
                for string in extract_strings_from_formula(str(value)):
                    if any(token in string for token in ("ОАЭ", "Турция", "Карго", "Польша", "Авто", "Авиадоставка", "сборный")):
                        all_delivery_strings.add(string)
                    if string.startswith("ИП"):
                        all_ip_strings.add(string)

        col_summary = {
            "column": letter,
            "index": col,
            "header": header,
            "sample_kind": cell_kind(sample_formula_value),
            "sample_value": sample_formula_value,
            "sample_cached_value": sample_cached_value,
            "counts": dict(kinds),
            "unique_inputs": unique_nonempty(unique_inputs, 50),
            "formula_templates": [
                {"formula": formula, "count": count}
                for formula, count in formula_templates.most_common(10)
            ],
        }
        columns.append(col_summary)

    delivery_values = []
    ip_values = []
    profit_values = []
    for row in range(first_data_row, last_data_row + 1):
        delivery_values.append(ws.cell(row, 4).value)  # D
        ip_values.append(ws.cell(row, 3).value)        # C
        profit_values.append(ws.cell(row, 40).value)   # AN

    aux_cells = []
    for row in range(1, aux.max_row + 1):
        row_values = []
        for col in range(1, aux.max_column + 1):
            value = aux.cell(row, col).value
            cached = aux_values.cell(row, col).value
            if value is not None or cached is not None:
                row_values.append({
                    "cell": f"{get_column_letter(col)}{row}",
                    "value": value,
                    "cached_value": cached,
                })
        if row_values:
            aux_cells.append({"row": row, "cells": row_values})

    data_validations = []
    try:
        for dv in ws.data_validations.dataValidation:
            data_validations.append({
                "type": dv.type,
                "sqref": str(dv.sqref),
                "formula1": dv.formula1,
                "formula2": dv.formula2,
            })
    except Exception as exc:  # pragma: no cover - defensive for Excel extensions
        data_validations.append({"error": str(exc)})

    settings_cells = {}
    for coord in ["I1", "I2", "K2", "AG4", "AO1", "AP1", "AT1", "AV1", "AQ4", "AR4", "BE6", "BF6", "BG6", "BH6", "BI6", "BJ6"]:
        settings_cells[coord] = {
            "formula": ws[coord].value,
            "cached_value": ws_values[coord].value,
        }

    summary = {
        "path": str(workbook_path),
        "main_sheet": "Калькулятор",
        "aux_sheet": "Сопутсвующие расходы",
        "data_area": f"A{first_data_row}:{get_column_letter(max_col)}{last_data_row}",
        "headers_row": header_row,
        "columns": columns,
        "delivery_values_from_column_D": unique_nonempty(delivery_values, 200),
        "ip_values_from_column_C": unique_nonempty(ip_values, 50),
        "profit_values_from_column_AN": unique_nonempty(profit_values, 50),
        "delivery_strings_from_formulas": sorted(all_delivery_strings),
        "ip_strings_from_formulas": sorted(all_ip_strings),
        "data_validations": data_validations,
        "settings_cells": settings_cells,
        "aux_cells": aux_cells,
    }

    json_path = out_dir / "volvo_calculator_logic.json"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    md_lines: list[str] = []
    md_lines.append("# Логика калькулятора Volvo")
    md_lines.append("")
    md_lines.append(f"Файл: `{workbook_path}`")
    md_lines.append(f"Основная область строк: `{summary['data_area']}`")
    md_lines.append("")
    md_lines.append("## Варианты")
    md_lines.append("")
    md_lines.append("Доставка из данных:")
    for value in summary["delivery_values_from_column_D"]:
        md_lines.append(f"- {value}")
    md_lines.append("")
    md_lines.append("Доставка из формул:")
    for value in summary["delivery_strings_from_formulas"]:
        md_lines.append(f"- {value}")
    md_lines.append("")
    md_lines.append("Выбор ИП:")
    for value in summary["ip_values_from_column_C"]:
        md_lines.append(f"- {value}")
    md_lines.append("")
    md_lines.append("Проценты прибыли из данных:")
    for value in summary["profit_values_from_column_AN"]:
        md_lines.append(f"- {value}")
    md_lines.append("")
    md_lines.append("## Настройки верхнего блока")
    md_lines.append("")
    md_lines.append("| Ячейка | Формула/значение | Кэш |")
    md_lines.append("|---|---|---|")
    for coord, item in settings_cells.items():
        md_lines.append(f"| {coord} | `{short(item['formula'])}` | `{short(item['cached_value'])}` |")
    md_lines.append("")
    md_lines.append("## Колонки строки калькулятора")
    md_lines.append("")
    md_lines.append("| Колонка | Заголовок | Тип в строке 6 | Пример | Кэш | Непустые входы | Формулы |")
    md_lines.append("|---|---|---|---|---|---:|---:|")
    for col in columns:
        counts = col["counts"]
        md_lines.append(
            f"| {col['column']} | {short(col['header'], 160)} | {col['sample_kind']} | "
            f"`{short(col['sample_value'], 220)}` | `{short(col['sample_cached_value'], 120)}` | "
            f"{counts.get('input', 0)} | {counts.get('formula', 0)} |"
        )
    md_lines.append("")
    md_lines.append("## Формулы по колонкам")
    for col in columns:
        if not col["formula_templates"]:
            continue
        md_lines.append("")
        md_lines.append(f"### {col['column']} {col['header']}")
        for item in col["formula_templates"][:5]:
            md_lines.append(f"- `{short(item['formula'], 900)}` ({item['count']} строк)")
    md_lines.append("")
    md_lines.append("## Таблица сопутствующих расходов")
    md_lines.append("")
    for row in aux_cells:
        parts = [f"{cell['cell']}=`{short(cell['value'], 120)}`" for cell in row["cells"]]
        md_lines.append(f"- row {row['row']}: " + "; ".join(parts))

    md_path = out_dir / "volvo_calculator_logic.md"
    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    print(json.dumps({
        "json": str(json_path),
        "markdown": str(md_path),
        "columns": len(columns),
        "delivery_options": len(summary["delivery_strings_from_formulas"]),
        "data_validations": len(data_validations),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
